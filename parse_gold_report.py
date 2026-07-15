"""
Gold Procurement TXT → Excel Converter
========================================
Parses Oracle fixed-width PO Transaction Reports and produces a clean
Excel with exactly 16 columns.

Usage:
    python parse_gold_report.py                          # all *.txt in parent dir
    python parse_gold_report.py "..\\FEB'25.txt"          # single file
    python parse_gold_report.py "..\\FEB'25.txt" "..\\Mar'25.txt"
"""

import sys
import re
import os
import glob
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Raw column definitions (must match separator groups in the Oracle report)
# ─────────────────────────────────────────────────────────────────────────────
RAW_COLUMNS = [
    "po_number", "po_date", "trans_date", "vendor_code", "vendor_name",
    "vendor_site_code", "state", "receipt_number", "receipt_date",
    "material_code", "material_description", "lot_number", "orgn_code",
    "subinv_code", "location", "prim_uom", "sec_uom",
    "primary_quantity", "secondary_quantity", "purchase_value",
    "gold_rate", "gold_net_weight", "gold_converted_weight", "gold_value",
    "stone_value", "loss_value", "labour_charges", "total_conv_weight",
    "total_gold_value", "tax_perc_1", "tax_perc_2", "tax_amount",
    "gstin", "glitem_class", "purity", "indicator", "other_mat_wt",
    "cfa_code", "attribute_category",
]

NUMERIC_COLS = {
    "primary_quantity", "secondary_quantity", "purchase_value",
    "gold_rate", "gold_net_weight", "gold_converted_weight",
    "gold_value", "stone_value", "loss_value", "labour_charges",
    "total_conv_weight", "total_gold_value",
    "tax_perc_1", "tax_perc_2", "tax_amount",
    "purity", "other_mat_wt",
}

# ─────────────────────────────────────────────────────────────────────────────
# Output columns  (the 16 columns you want in the Excel)
# ─────────────────────────────────────────────────────────────────────────────
OUTPUT_HEADERS = [
    "Source Type",
    "P.O.Number",
    "P.O.Date",
    "Name of the Supplier",
    "Delivery Location",
    "Loan Type",
    "Qty",
    "Qty - 995",
    "Purity - %",
    "Int",
    "Pre",
    "Duty",
    "Qtr",
    "Month",
    "TOTAL VAL EX GST",
    "PRICE EX GST",
]

OUTPUT_FIELDS = [
    "source_type",
    "po_number",
    "po_date",
    "vendor_name",
    "location",
    "loan_type",
    "qty",
    "qty_995",
    "purity",
    "int_val",
    "pre_val",
    "duty_val",
    "qtr",
    "month",
    "total_val_ex_gst",
    "price_ex_gst",
]

# ─────────────────────────────────────────────────────────────────────────────
# Noise-line patterns
# ─────────────────────────────────────────────────────────────────────────────
_RE_PAGE_HEADER  = re.compile(
    r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)',
    re.IGNORECASE,
)
_RE_SEPARATOR    = re.compile(r'^-{5,}')
_RE_REPORT_TITLE = re.compile(r'TJD CST PO TRANSACTIONS REPORT', re.IGNORECASE)
_RE_NON_HDR      = re.compile(r'^\s*(Non|Gold\s+Rec\s+Recover|other)\s*$', re.IGNORECASE)
_RE_DATA_LINE    = re.compile(r'^\d{7}\s')

# Delivery location mapping  (subinv_code → city)
_LOCATION_MAP = {
    "PRRM-ERRM": "KOLKATA",
    "RM-WSRM":   "MUMBAI",
    "RM-HSRM":   "HOSUR",
    "RM-N1RM":   "DELHI",
}


# ─────────────────────────────────────────────────────────────────────────────
# Helper: derive column char-position ranges from the separator line
# ─────────────────────────────────────────────────────────────────────────────
def _get_colspecs(separator_line: str) -> list[tuple[int, int]]:
    return [(m.start(), m.end()) for m in re.finditer(r'-+', separator_line)]


# ─────────────────────────────────────────────────────────────────────────────
# Source Type  — derived from material_code
# ─────────────────────────────────────────────────────────────────────────────
def _source_type(mat_code: str) -> str:
    mc = mat_code.strip().upper()
    if mc.startswith("11GORYM"):
        return "Bank"
    if mc.startswith("11GOZ") or mc.startswith("11GOP"):
        return "Exchange"
    if mc.startswith("11GOR"):
        return "Bank"
    if mc.startswith("11PAZ"):
        return "Alloy"
    if mc.startswith("11COZ"):
        return "Copper"
    return "Other"


# ─────────────────────────────────────────────────────────────────────────────
# Loan Type  — derived from attribute_category / glitem_class
# ─────────────────────────────────────────────────────────────────────────────
def _loan_type(row: dict) -> str:
    attr = str(row.get("attribute_category", "")).strip().upper()
    gl   = str(row.get("glitem_class", "")).strip().upper()
    lot  = str(row.get("lot_number", "")).strip().upper()
    if attr == "EXCHANGE":
        return "Exchange"
    if lot == "2B" and gl == "PRECMATL":
        return "Gold Metal Loan"
    if gl == "RAW MATL":
        return "Outright"
    if gl == "PRECMATL":
        return "Gold Metal Loan"
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Quarter  — Indian Financial Year  (Q4 = Jan-Mar, Q1 = Apr-Jun, …)
# ─────────────────────────────────────────────────────────────────────────────
_MONTH_TO_QTR = {
    "JAN": "Q4", "FEB": "Q4", "MAR": "Q4",
    "APR": "Q1", "MAY": "Q1", "JUN": "Q1",
    "JUL": "Q2", "AUG": "Q2", "SEP": "Q2",
    "OCT": "Q3", "NOV": "Q3", "DEC": "Q3",
}


def _quarter_from_date(po_date: str) -> str:
    """po_date looks like '08-FEB-25'."""
    m = re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)',
                  po_date, re.IGNORECASE)
    if m:
        return _MONTH_TO_QTR.get(m.group(1).upper(), "")
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Core: parse one TXT file → DataFrame  (only the 16 output columns)
# ─────────────────────────────────────────────────────────────────────────────
def parse_txt(filepath: str) -> pd.DataFrame:
    path = Path(filepath)
    print(f"\n{'─' * 60}")
    print(f"  Parsing: {path.name}")
    print(f"{'─' * 60}")

    with open(filepath, encoding="utf-8", errors="replace") as fh:
        lines = fh.readlines()

    colspecs   = None
    data_lines = []

    for raw in lines:
        line = raw.rstrip("\n")
        if not line.strip():
            continue
        if _RE_PAGE_HEADER.match(line):
            continue
        if _RE_REPORT_TITLE.search(line):
            continue
        if _RE_NON_HDR.match(line):
            continue
        if line.lstrip().startswith("PO") or line.lstrip().startswith("Number"):
            continue
        if _RE_SEPARATOR.match(line.lstrip()):
            if colspecs is None:
                colspecs = _get_colspecs(line)
            continue
        if _RE_DATA_LINE.match(line.strip()):
            data_lines.append(line)

    if colspecs is None:
        raise ValueError(f"Could not find separator line in {filepath}")

    expected = len(RAW_COLUMNS)
    if len(colspecs) < expected:
        last_end = colspecs[-1][1]
        while len(colspecs) < expected:
            s = last_end + 1
            colspecs.append((s, s + 30))
            last_end = s + 30
    colspecs = colspecs[:expected]

    print(f"  Detected {len(colspecs)} columns | {len(data_lines)} data lines")

    # ── Parse raw rows ──
    raw_records = []
    for line in data_lines:
        padded = line.ljust(colspecs[-1][1] + 5)
        row = {}
        for fname, (start, end) in zip(RAW_COLUMNS, colspecs):
            val = padded[start:end].strip()
            if fname in NUMERIC_COLS:
                val = val.lstrip("$").replace(",", "").strip()
                try:
                    row[fname] = float(val) if val and val != "." else 0.0
                except ValueError:
                    row[fname] = 0.0
            else:
                row[fname] = val
        raw_records.append(row)

    # ── Month label from filename (e.g. "FEB'25" → "FEB-25") ──
    month_label = re.sub(r"['\s]", "-", path.stem).upper()

    # ── Filter: only material codes starting with 11GORY ──
    raw_records = [r for r in raw_records
                   if r["material_code"].strip().upper().startswith("11GORY")]

    # ── Filter: exclude WATCH DIVISION from vendor name ──
    raw_records = [r for r in raw_records
                   if "WATCH DIVISION" not in r["vendor_name"].strip().upper()]

    # ── Build output records with only the 16 columns ──
    out_records = []
    for r in raw_records:
        qty    = r["primary_quantity"]
        purity = r["purity"]
        val    = r["purchase_value"]

        qty_995  = (qty * purity / 995) if purity else 0.0
        price_ex = val / qty if qty else 0.0

        out_records.append({
            "source_type":      _source_type(r["material_code"]),
            "loan_type":        "",
            "po_number":        r["po_number"],
            "po_date":          r["po_date"],
            "vendor_name":      r["vendor_name"],
            "location":         _LOCATION_MAP.get(r["location"].strip().upper(), r["location"]),
            "qty":              qty,
            "qty_995":          round(qty_995, 4),
            "purity":           purity,
            "int_val":          "",
            "pre_val":          "",
            "duty_val":         "",
            "qtr":              _quarter_from_date(r["po_date"]),
            "month":            month_label,
            "total_val_ex_gst": val,
            "price_ex_gst":     round(price_ex, 2),
        })

    df = pd.DataFrame(out_records, columns=OUTPUT_FIELDS)

    print(f"  Rows               : {len(df)}")
    print(f"  Total Qty          : {df['qty'].sum():,.3f} GMS")
    print(f"  Total Val Ex GST   : ₹{df['total_val_ex_gst'].sum():,.2f}")
    print(f"  Unique suppliers   : {df['vendor_name'].nunique()}")

    return df


# ─────────────────────────────────────────────────────────────────────────────
# Excel styling
# ─────────────────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="006966")
_ALT_FILL    = PatternFill("solid", fgColor="ECF7F6")
_TOTAL_FILL  = PatternFill("solid", fgColor="FFF9C4")
_THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

_NUMERIC_FMT = {
    "qty":              "#,##0.000",
    "qty_995":          "#,##0.0000",
    "purity":           "0.00",
    "total_val_ex_gst": "#,##0.00",
    "price_ex_gst":     "#,##0.00",
}


def _autofit(ws, min_w=10, max_w=35):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = min_w
        for cell in col:
            try:
                clen = len(str(cell.value)) if cell.value else 0
                if clen > best:
                    best = clen
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(best + 3, max_w)


def _write_sheet(wb, df: pd.DataFrame, sheet_name: str = "Gold Procurement"):
    ws = wb.create_sheet(sheet_name)

    # Header row
    ws.append(OUTPUT_HEADERS)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

    # Data rows
    records = df.to_dict("records")
    for row_idx, rec in enumerate(records, start=2):
        row_data = [rec.get(k, "") for k in OUTPUT_FIELDS]
        ws.append(row_data)
        for cell, fkey in zip(ws[row_idx], OUTPUT_FIELDS):
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = _ALT_FILL
            if fkey in _NUMERIC_FMT:
                cell.number_format = _NUMERIC_FMT[fkey]

    # Totals row
    last_data = ws.max_row
    ws.append(["TOTAL"] + [""] * (len(OUTPUT_HEADERS) - 1))
    tr = ws[ws.max_row]
    tr[0].font = Font(bold=True)
    tr[0].fill = _TOTAL_FILL

    for fkey in ("qty", "qty_995", "total_val_ex_gst"):
        col_idx = OUTPUT_FIELDS.index(fkey) + 1
        col_ltr = get_column_letter(col_idx)
        cell = tr[col_idx - 1]
        cell.value         = f"=SUM({col_ltr}2:{col_ltr}{last_data})"
        cell.font          = Font(bold=True)
        cell.fill          = _TOTAL_FILL
        cell.number_format = _NUMERIC_FMT.get(fkey, "#,##0.00")
        cell.border        = _THIN_BORDER

    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(df: pd.DataFrame, out_path: str):
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    _write_sheet(wb, df)
    wb.save(out_path)
    print(f"\n  ✓ Saved: {out_path}")


def write_excel_to_buffer(df: pd.DataFrame) -> tuple:
    """Return (bytes, stats_dict) — used by Flask web API."""
    import io
    stats = {
        "rows":            len(df),
        "purchase_value":  round(df["total_val_ex_gst"].sum(), 2),
        "tax_amount":      0,
        "qty_gms":         round(df["qty"].sum(), 3),
        "unique_vendors":  int(df["vendor_name"].nunique()),
        "month":           df["month"].iloc[0] if len(df) else "",
        "vendor_types":    df["source_type"].value_counts().to_dict(),
    }
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    _write_sheet(wb, df)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), stats


# ─────────────────────────────────────────────────────────────────────────────
# Report sheet  — Procurement loan-type categories (team fills Loan Type in Master)
# ─────────────────────────────────────────────────────────────────────────────
PROCUREMENT_LOAN_TYPES = [
    "SBLC", "CLEAN GDS", "CLEAN", "GOL DOM",
    "EXPORT CLEAN", "EXPORT SBLC", "EXPORT",
    "GOL", "SPOT CEPA", "SPOT IIBX", "SPOT FWD", "SPOT HUTTI GOLD", "SPOT",
]

_RPT_TITLE_F = PatternFill("solid", fgColor="1B4F72")
_RPT_HDR_F   = PatternFill("solid", fgColor="2980B9")
_RPT_MAN_F   = PatternFill("solid", fgColor="FEF9E7")
_RPT_ACT_F   = PatternFill("solid", fgColor="EBF5FB")
_RPT_FRM_F   = PatternFill("solid", fgColor="EAFAF1")
_RPT_TOT_F   = PatternFill("solid", fgColor="FFF9C4")
_RPT_ALT_F   = PatternFill("solid", fgColor="F5F5F5")
_RPT_THICK_B = Border(
    left=Side(style="medium", color="2980B9"),
    right=Side(style="medium", color="2980B9"),
    top=Side(style="medium", color="2980B9"),
    bottom=Side(style="medium", color="2980B9"),
)


def _write_report_sheet(wb: Workbook, master_df: pd.DataFrame):
    """
    Build / replace the 'Report' sheet in the master workbook.

    Sections
    --------
    1. Budget vs Actual Qty (KGs) – Budget: manual | Actual: from Qty col
    2. GOL Interest Rate          – Budget: manual | Actual: from Int col
    3. Premium Budget vs Actual   – Budget: manual | Actual: from Pre col
    4. Gold Input Alignment       – alignment %: manual | Qty: from Qty col
    5. Procurement Details Qty    – grouped by Loan Type (team fills)
    6. Procurement Details Value  – grouped by Loan Type (team fills)
    """
    if "Report" in wb.sheetnames:
        del wb["Report"]
    ws = wb.create_sheet("Report")

    months = sorted(master_df["month"].unique().tolist())
    n      = len(months)

    # ── Aggregate data by month ──────────────────────────────────────────
    qty_m = (
        master_df.groupby("month")["qty"].sum()
        .div(1000).reindex(months, fill_value=0)      # grams → KGs
    )
    int_m = (
        master_df.assign(
            int_val=pd.to_numeric(master_df["int_val"], errors="coerce").fillna(0)
        ).groupby("month")["int_val"].sum().reindex(months, fill_value=0)
    )
    pre_m = (
        master_df.assign(
            pre_val=pd.to_numeric(master_df["pre_val"], errors="coerce").fillna(0)
        ).groupby("month")["pre_val"].sum().reindex(months, fill_value=0)
    )
    val_m = (
        master_df.groupby("month")["total_val_ex_gst"].sum()
        .reindex(months, fill_value=0)
    )

    # Procurement pivots (loan_type × month) — uppercase for case-insensitive match
    lt_df = master_df.copy()
    lt_df["loan_type"] = lt_df["loan_type"].fillna("").str.strip().str.upper()
    qty_lt = (
        lt_df.groupby(["loan_type", "month"])["qty"].sum()
        .unstack(fill_value=0).div(1000)
        .reindex(columns=months, fill_value=0)
    )
    val_lt = (
        lt_df.groupby(["loan_type", "month"])["total_val_ex_gst"].sum()
        .unstack(fill_value=0).reindex(columns=months, fill_value=0)
    )

    grand_qty = float(qty_m.sum())
    grand_val = float(val_m.sum())

    # ── Style / alignment helpers ──────────────────────────────────────────
    _TB  = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD"),
    )
    _CTR = Alignment(horizontal="center", vertical="center")
    _LFT = Alignment(horizontal="left",   vertical="center", indent=1)
    _LF0 = Alignment(horizontal="left",   vertical="center")
    _RGT = Alignment(horizontal="right",  vertical="center")

    cur = [1]      # mutable row pointer
    def R():    return cur[0]
    def adv():  cur[0] += 1

    def W(r, c, v=None, fill=None, font=None, align=None, fmt=None, bdr=None):
        cell = ws.cell(row=r, column=c, value=v)
        if fill:  cell.fill          = fill
        if font:  cell.font          = font
        if align: cell.alignment     = align
        if fmt:   cell.number_format = fmt
        cell.border = bdr if bdr else _TB
        return cell

    # ── Section title (full-width merged) ───────────────────────────────
    def sec(text, ncols):
        r = R()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        W(r, 1, text,
          fill=_RPT_TITLE_F,
          font=Font(bold=True, color="FFFFFF", size=11),
          align=_CTR, bdr=_RPT_THICK_B)
        ws.row_dimensions[r].height = 24
        adv()

    # ── Column header row ─────────────────────────────────────────────
    def hdr(first, share=False):
        r = R()
        cols = [first] + months + ["Total"] + (["% Share"] if share else [])
        for ci, h in enumerate(cols, 1):
            W(r, ci, h,
              fill=_RPT_HDR_F,
              font=Font(bold=True, color="FFFFFF", size=10),
              align=_CTR if ci > 1 else _LF0,
              bdr=_RPT_THICK_B)
        ws.row_dimensions[r].height = 18
        adv()

    # ── Manual (blank) row — team fills ─────────────────────────────────
    def man(label, fmt="#,##0.000"):
        r = R()
        W(r, 1, label,
          fill=_RPT_MAN_F,
          font=Font(size=10, italic=True, color="808080"),
          align=_LFT)
        for ci in range(2, n + 3):
            W(r, ci, None, fill=_RPT_MAN_F, fmt=fmt)
        adv()
        return r

    # ── Actual (auto-populated) row ─────────────────────────────────
    def act(label, vals, fmt="#,##0.000"):
        r = R()
        total = sum(float(v) for v in vals)
        W(r, 1, label, fill=_RPT_ACT_F,
          font=Font(size=10, bold=True), align=_LFT)
        for ci, v in enumerate(vals, 2):
            W(r, ci, round(float(v), 6), fill=_RPT_ACT_F,
              font=Font(size=10), align=_RGT, fmt=fmt)
        W(r, n + 2, round(total, 6), fill=_RPT_ACT_F,
          font=Font(size=10, bold=True), align=_RGT, fmt=fmt)
        adv()
        return r

    # ── Formula row ──────────────────────────────────────────────────
    def fml(label, bgt_r, act_r, kind="diff", fmt="#,##0.000"):
        """kind: 'pct' = (actual/budget)*100  |  'diff' = actual-budget"""
        r = R()
        W(r, 1, label, fill=_RPT_FRM_F,
          font=Font(size=10, italic=True), align=_LFT)
        for ci in range(2, n + 3):          # month cols + Total
            col = get_column_letter(ci)
            if kind == "pct":
                formula = f'=IFERROR({col}{act_r}/{col}{bgt_r}*100,"")'
                num     = "0.00"
            else:
                formula = f"={col}{act_r}-{col}{bgt_r}"
                num     = fmt
            W(r, ci, formula, fill=_RPT_FRM_F,
              font=Font(size=10), align=_RGT, fmt=num)
        adv()

    # ── Procurement data row ──────────────────────────────────────────
    def prc(label, vals, fmt="#,##0.000",
            is_total=False, alt=False, grand=None):
        r = R()
        total = sum(float(v) for v in vals)
        fill  = _RPT_TOT_F if is_total else (_RPT_ALT_F if alt else None)
        font  = Font(bold=True, size=10) if is_total else Font(size=10)
        W(r, 1, label, fill=fill, font=font,
          align=_LF0 if is_total else _LFT)
        for ci, v in enumerate(vals, 2):
            W(r, ci, round(float(v), 4), fill=fill, font=font,
              align=_RGT, fmt=fmt)
        W(r, n + 2, round(total, 4), fill=fill, font=font,
          align=_RGT, fmt=fmt)
        if grand is not None:
            pct = (100.0 if is_total
                   else (round(total / grand * 100, 2) if grand else 0))
            W(r, n + 3, pct, fill=fill, font=font, align=_RGT, fmt="0.00")
        adv()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 1 — Budget vs Actual (Qty, KGs)
    # ══════════════════════════════════════════════════════════════════════
    sec("Gold Procurement: Budget Vs Actual  [Qty in KGs]", n + 2)
    hdr("Description")
    bgt1 = man("Budgeted Qty in KGs")
    act1 = act("Actual Qty in KGs", qty_m.tolist())
    fml("Achievement %",                      bgt1, act1, kind="pct")
    fml("Decrease / Increase w.r.t Budget",   bgt1, act1, kind="diff")
    adv()   # blank separator

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 2 — GOL Interest Rate
    # ══════════════════════════════════════════════════════════════════════
    sec("GOL - Interest Rate (including Export)", n + 2)
    hdr("Details")
    bgt2 = man("Budget", fmt="0.0000")
    act2 = act("Actual",  int_m.tolist(), fmt="0.0000")
    fml("Variation", bgt2, act2, kind="diff", fmt="0.0000")
    adv()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 3 — Premium Budget vs Actual
    # ══════════════════════════════════════════════════════════════════════
    sec("Premium: Budget Vs Actual  [$ / T.oz]", n + 2)
    hdr("Details")
    bgt3 = man("Budgeted Premium - $ / T.oz", fmt="0.0000")
    act3 = act("Actual Premium - $ / T.oz",   pre_m.tolist(), fmt="0.0000")
    fml("Variation - $ / T.oz", bgt3, act3, kind="diff", fmt="0.0000")
    adv()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 4 — Gold Input Alignment
    # ══════════════════════════════════════════════════════════════════════
    sec("Gold Input Alignment - including Export (mixed purity)", n + 2)
    hdr("Details")
    man("Gold delivery alignment in %", fmt="0.00%")
    act("Quantity purchased in KGs",    qty_m.tolist())
    adv()
    adv()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 5 — Procurement Details (Qty, KGs)
    # ══════════════════════════════════════════════════════════════════════
    sec("Procurement Details (mixed purity)  [Qty in KGs]", n + 3)
    hdr("Details", share=True)
    for idx, lt in enumerate(PROCUREMENT_LOAN_TYPES):
        key  = lt.upper()
        vals = qty_lt.loc[key].tolist() if key in qty_lt.index else [0.0] * n
        prc(lt, vals, alt=(idx % 2 == 1), grand=grand_qty)
    prc("Total Qty - KGs", qty_m.tolist(), is_total=True, grand=grand_qty)
    adv()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 6 — Procurement Details (Value, ₹)
    # ══════════════════════════════════════════════════════════════════════
    sec("Procurement Details - Value  [\u20b9 Ex GST]", n + 3)
    hdr("Details", share=True)
    for idx, lt in enumerate(PROCUREMENT_LOAN_TYPES):
        key  = lt.upper()
        vals = val_lt.loc[key].tolist() if key in val_lt.index else [0.0] * n
        prc(lt, vals, fmt="#,##0.00", alt=(idx % 2 == 1), grand=grand_val)
    prc("Total Value - \u20b9", val_m.tolist(),
        fmt="#,##0.00", is_total=True, grand=grand_val)

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 45
    for ci in range(2, n + 4):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# Master Excel  — accumulates all months in a single workbook
# ─────────────────────────────────────────────────────────────────────────────
MASTER_FILENAME = "GoldProcurement_Master.xlsx"


def _get_master_path(out_dir: str | Path) -> Path:
    """Return the path to the master Excel in the given directory."""
    return Path(out_dir) / MASTER_FILENAME


def update_master_excel(df: pd.DataFrame, out_dir: str | Path):
    """
    Append the current month's data to the master Excel.
    - If the master doesn't exist, create it.
    - If the month already exists in the master, replace that month's data.
    - The master has one sheet with ALL months combined.
    """
    master_path = _get_master_path(out_dir)

    # Determine month label from the dataframe
    if df.empty:
        return
    current_month = df["month"].iloc[0]

    # Load existing master data (if any)
    if master_path.exists():
        existing_df = pd.read_excel(master_path, sheet_name="Master",
                                    dtype=str, header=0)
        # Map Excel headers back to internal field names
        header_to_field = dict(zip(OUTPUT_HEADERS, OUTPUT_FIELDS))
        existing_df.columns = [header_to_field.get(c, c) for c in existing_df.columns]

        # Drop the TOTAL summary row that _write_sheet appends
        existing_df = existing_df[
            existing_df[existing_df.columns[0]].astype(str).str.upper() != "TOTAL"
        ]

        # Convert numeric columns back; string cols: NaN → ""
        for col in ("qty", "qty_995", "purity", "total_val_ex_gst", "price_ex_gst",
                    "int_val", "pre_val", "duty_val"):
            if col in existing_df.columns:
                existing_df[col] = pd.to_numeric(existing_df[col], errors="coerce").fillna(0.0)
        for col in ("source_type", "loan_type"):
            if col in existing_df.columns:
                existing_df[col] = existing_df[col].fillna("")

        # If this month already exists in the master, skip — do not overwrite
        if current_month in existing_df["month"].values:
            print(f"  ⏭ Master already has {current_month} — skipping (no overwrite)")
            return

        # Append new month's data
        master_df = pd.concat([existing_df, df], ignore_index=True)
    else:
        master_df = df.copy()

    # Sort by month for consistent ordering
    master_df = master_df.reset_index(drop=True)

    # Write master workbook (Master data sheet + Report summary sheet)
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    _write_sheet(wb, master_df, sheet_name="Master")
    _write_report_sheet(wb, master_df)
    wb.save(str(master_path))
    print(f"  ✓ Master updated: {master_path}  ({len(master_df)} total rows, Report refreshed)")


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────
def _infer_month_tag(filepath: str) -> str:
    stem = Path(filepath).stem
    return re.sub(r"['\s]+", "_", stem).upper()


def process_file(filepath: str):
    if not os.path.exists(filepath):
        print(f"  ✗ File not found: {filepath}")
        return
    df = parse_txt(filepath)
    if df.empty:
        print(f"  ✗ No data rows parsed from {filepath}")
        return
    tag      = _infer_month_tag(filepath)
    out_dir  = Path(filepath).parent
    out_path = str(out_dir / f"GoldProcurement_{tag}.xlsx")
    write_excel(df, out_path)

    # Update master Excel with this month's data
    update_master_excel(df, out_dir)


def main():
    if len(sys.argv) > 1:
        files = sys.argv[1:]
    else:
        # Look in parent directory for .txt files
        parent = str(Path(__file__).resolve().parent.parent)
        files  = glob.glob(os.path.join(parent, "*.txt"))
        if not files:
            files = glob.glob("*.txt")
        if not files:
            print("No .txt files found.")
            print("Usage: python parse_gold_report.py <file1.txt> [file2.txt ...]")
            sys.exit(1)
        print(f"Auto-discovered {len(files)} file(s): "
              f"{', '.join(Path(f).name for f in files)}")

    for f in files:
        try:
            process_file(f)
        except Exception as exc:
            print(f"\n  ✗ Error processing {f}: {exc}")
            import traceback
            traceback.print_exc()

    print("\nDone.")


if __name__ == "__main__":
    main()
